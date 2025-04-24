import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from datetime import datetime

def process_excel(file_stream):
    df = pd.read_excel(file_stream)

    df = df[df['task_status'].str.strip().str.lower() == 'processing']

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    status_options = [
        "Come on Saturday", "Stay late", "Come Early", "Task reassigned", "rejected",
        "Inprocess", "Convert to workflow", "On hold (Specific Reason)", "Completed",
        "Pending", "QC (Specific Reason)", "Other (specific reason)", "Decision pending",
        "Cascading Task", "Escalated", "Workflow"
    ]
    reason_options = [
        "none", "Employee > can't complete this task as the task communication is not ended",
        "Employee > come Early", "Employ > Come Saturday",
        "Employee > I have completed this task & pending from other task member",
        "Employee > I have completed this task & pending from the task assigner",
        "Employee > Salary Deduction", "Employee > Stay Late",
        "Management > I have completed this task & it's pending from other task member who is in the management",
        "Management > I have completed this task and it's pending from other task member who is in no longer in the company",
        "Management > I completed this task & it's pending from task assigner who is in the management",
        "Management > task waiting for a client approval", "QC (Specify Reason", "On hold(Specific Reason)",
        "Other (Specify Reason)", "Adhoc task"
    ]
    priority_colors = {
        "High": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
        "Urgent": PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid"),
        "SOS": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
        "Normal": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    }

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    date_format = 'DD-MMM-YYYY'

    for assignee in df['task_assignee'].unique():
        user_df = df[df['task_assignee'] == assignee]
        sheet = wb.create_sheet(title=assignee[:31])

        headers = ["Sr No", "Week", "Meeting Date", "Task Id", "Task Title", "Task Priority",
                   "Task Status", "Actual Start Date", "Assign Date", "Due Date",
                   "Actual End Date", "Reason", "Status", "Discussion Point"]
        sheet.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        for row_idx, (_, row) in enumerate(user_df.iterrows(), start=2):
            sheet.append([
                row_idx - 1, "", "", row['master_task_id'], row['master_task_title'],
                row.get('task_priority', ""), row['task_status'],
                row['master_task_update_date'], row['master_task_start_date'],
                row['master_task_tentative_end_date'], row['master_task_actual_end_date'],
                "", "", ""
            ])

        max_row = sheet.max_row
        reason_dv = DataValidation(type="list", formula1='"' + ','.join(reason_options) + '"', allow_blank=True)
        status_dv = DataValidation(type="list", formula1='"' + ','.join(status_options) + '"', allow_blank=True)
        sheet.add_data_validation(reason_dv)
        sheet.add_data_validation(status_dv)

        for row in range(2, max_row + 1):
            reason_dv.add(sheet[f'L{row}'])
            status_dv.add(sheet[f'M{row}'])
            for col in range(1, 15):
                sheet.cell(row=row, column=col).alignment = align_center
            for col_letter in ['H', 'I', 'J']:
                cell = sheet[f'{col_letter}{row}']
                try:
                    if cell.value:
                        cell.value = pd.to_datetime(cell.value).date()
                        cell.number_format = date_format
                except:
                    pass
            cell = sheet[f'F{row}']
            value = str(cell.value).strip()
            if value in priority_colors:
                cell.fill = priority_colors[value]
                cell.font = Font(color="000000")

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, df
